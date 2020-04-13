import sys, openpyxl
from openpyxl.styles import Font

n = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb['Sheet']

# label
for i in range(2, n+2):
    # bold font
    bold = Font(bold=True)
    # column label
    sheet.cell(row=1, column=i).value = i - 1
    sheet.cell(row=1, column=i).font = bold
    # row label
    sheet.cell(row=i, column=1).value = i - 1
    sheet.cell(row=i, column=1).font = bold

# content
for i in range(2, n+2):
    for j in range(2, n+2):
        sheet.cell(row=i, column=j).value = (i - 1) * (j - 1)

wb.save('multiplicationTable.xlsx')